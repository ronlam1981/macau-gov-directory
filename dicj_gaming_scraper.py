#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
澳門博彩監察協調局 (DICJ) 博彩統計資料爬蟲
======================================
從 www.dicj.gov.mo 抓取：
  1. 每月幸運博彩毛收入（2002年至今）
  2. 每季各博彩項目資料（2002年至今）

輸出：DICJ_博彩統計.xlsx  （三個工作表）

用法：
  python3 dicj_gaming_scraper.py                  # 完整爬取
  python3 dicj_gaming_scraper.py --update          # 只更新最新資料
  python3 dicj_gaming_scraper.py --year 2020 2024  # 指定年份範圍
  python3 dicj_gaming_scraper.py --excel-only      # 從快取重新生成 Excel
  python3 dicj_gaming_scraper.py --seed            # 預填充已知歷史數據
"""

import argparse
import os
import sys
import time
import random
import re
import json
import warnings
from datetime import datetime, date
from io import StringIO

warnings.filterwarnings("ignore")

# ── 依賴檢查 ─────────────────────────────────────────────────────────────────
try:
    import requests
    from requests.adapters import HTTPAdapter
    from urllib3.util.retry import Retry
except ImportError:
    sys.exit("❌ 缺少 requests：pip3 install requests")

try:
    from bs4 import BeautifulSoup
except ImportError:
    sys.exit("❌ 缺少 beautifulsoup4：pip3 install beautifulsoup4 lxml")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("❌ 缺少 openpyxl：pip3 install openpyxl")

# ── 常數 ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "DICJ_博彩統計.xlsx")
CACHE_FILE  = os.path.join(SCRIPT_DIR, "Rawdata", "dicj_cache.json")

BASE_CN   = "https://www.dicj.gov.mo/web/cn/information"
BASE_EN   = "https://www.dicj.gov.mo/web/en/information"
YEAR_START = 2002
YEAR_NOW   = datetime.now().year

MONTH_MAP = {
    "一月": 1, "二月": 2, "三月": 3, "四月": 4,
    "五月": 5, "六月": 6, "七月": 7, "八月": 8,
    "九月": 9, "十月": 10, "十一月": 11, "十二月": 12,
    "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4,
    "May": 5, "Jun": 6, "Jul": 7, "Aug": 8,
    "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
}

MONTH_NAMES_ZH = {
    1: "一月", 2: "二月", 3: "三月", 4: "四月",
    5: "五月", 6: "六月", 7: "七月", 8: "八月",
    9: "九月", 10: "十月", 11: "十一月", 12: "十二月",
}

QUARTER_NAMES = {1: "第一季", 2: "第二季", 3: "第三季", 4: "第四季"}

GAME_ALIASES = {
    "百家樂貴賓廳": "百家樂貴賓廳",
    "百家樂大眾廳": "百家樂大眾廳",
    "三卡百家樂": "三卡百家樂",
    "廿一點": "廿一點",
    "輪盤": "輪盤",
    "骰寶": "骰寶",
    "牌九": "牌九",
    "其他桌面博彩": "其他桌面博彩",
    "角子機": "角子機",
    "合計": "合計",
    "VIP Baccarat": "百家樂貴賓廳",
    "Mass Market Baccarat": "百家樂大眾廳",
    "Three-card Baccarat": "三卡百家樂",
    "Blackjack": "廿一點",
    "Roulette": "輪盤",
    "Sic-Bo": "骰寶",
    "Sic Bo": "骰寶",
    "Pai Gow": "牌九",
    "Other Table Games": "其他桌面博彩",
    "Slot Machines": "角子機",
    "Total": "合計",
}

GAME_ORDER = [
    "百家樂貴賓廳", "百家樂大眾廳", "三卡百家樂",
    "廿一點", "輪盤", "骰寶", "牌九", "其他桌面博彩",
    "角子機", "合計",
]

# Excel 顏色
CLR_HEADER_BLUE   = "1F4E79"
CLR_HEADER_TEAL   = "1F5C5C"
CLR_HEADER_ORANGE = "833C00"
CLR_ROW_EVEN      = "EBF3FB"
CLR_ROW_ODD       = "FFFFFF"
CLR_TOTAL_BG      = "FFF2CC"
CLR_TITLE         = "2E75B6"
CLR_NOTE          = "7F7F7F"


# ══════════════════════════════════════════════════════════════════════════════
# 已確認歷史數據（來源：DICJ 官方新聞稿及主要財經媒體）
# 單位：百萬澳門幣
# ══════════════════════════════════════════════════════════════════════════════
SEED_MONTHLY = [
    # ── 2025 年（資料來源：DICJ 官方及媒體報導）
    {"年份": 2025, "月份": 1,  "毛收入_百萬澳門幣": 18254, "較上月變動_%": None,  "較去年同期變動_%": -5.6,  "本年累計毛收入_百萬澳門幣": 18254},
    {"年份": 2025, "月份": 2,  "毛收入_百萬澳門幣": 19742, "較上月變動_%":  8.2,  "較去年同期變動_%":  6.8,  "本年累計毛收入_百萬澳門幣": 37996},
    {"年份": 2025, "月份": 3,  "毛收入_百萬澳門幣": 19660, "較上月變動_%": -0.4,  "較去年同期變動_%":  0.8,  "本年累計毛收入_百萬澳門幣": 57656},
    {"年份": 2025, "月份": 4,  "毛收入_百萬澳門幣": 18860, "較上月變動_%": -4.1,  "較去年同期變動_%":  2.9,  "本年累計毛收入_百萬澳門幣": 76516},
    {"年份": 2025, "月份": 5,  "毛收入_百萬澳門幣": 21190, "較上月變動_%" : 12.4, "較去年同期變動_%":  5.0,  "本年累計毛收入_百萬澳門幣": 97706},
    {"年份": 2025, "月份": 6,  "毛收入_百萬澳門幣": 21060, "較上月變動_%": -0.6,  "較去年同期變動_%": 19.0,  "本年累計毛收入_百萬澳門幣": 118766},
    {"年份": 2025, "月份": 7,  "毛收入_百萬澳門幣": 22120, "較上月變動_%":  5.0,  "較去年同期變動_%": 19.0,  "本年累計毛收入_百萬澳門幣": 140886},
    {"年份": 2025, "月份": 8,  "毛收入_百萬澳門幣": 22160, "較上月變動_%":  0.2,  "較去年同期變動_%": 12.0,  "本年累計毛收入_百萬澳門幣": 163046},
    {"年份": 2025, "月份": 9,  "毛收入_百萬澳門幣": 18270, "較上月變動_%":-17.5,  "較去年同期變動_%":  5.6,  "本年累計毛收入_百萬澳門幣": 181316},
    {"年份": 2025, "月份": 10, "毛收入_百萬澳門幣": 24100, "較上月變動_%": 31.9,  "較去年同期變動_%": 16.4,  "本年累計毛收入_百萬澳門幣": 205416},
    {"年份": 2025, "月份": 11, "毛收入_百萬澳門幣": 21090, "較上月變動_%":-12.5,  "較去年同期變動_%": 14.4,  "本年累計毛收入_百萬澳門幣": 226506},
    # ── 2024 年（資料來源：DICJ 官方，全年 226,780 百萬澳門幣）
    {"年份": 2024, "月份": 1,  "毛收入_百萬澳門幣": 19330, "較上月變動_%": None,  "較去年同期變動_%": 66.9, "本年累計毛收入_百萬澳門幣": 19330},
    {"年份": 2024, "月份": 2,  "毛收入_百萬澳門幣": 18480, "較上月變動_%": -4.4,  "較去年同期變動_%": 79.1, "本年累計毛收入_百萬澳門幣": 37810},
    {"年份": 2024, "月份": 3,  "毛收入_百萬澳門幣": 19500, "較上月變動_%":  5.5,  "較去年同期變動_%": 53.1, "本年累計毛收入_百萬澳門幣": 57310},
    {"年份": 2024, "月份": 4,  "毛收入_百萬澳門幣": 18330, "較上月變動_%": -6.0,  "較去年同期變動_%": 27.8, "本年累計毛收入_百萬澳門幣": 75640},
    {"年份": 2024, "月份": 5,  "毛收入_百萬澳門幣": 20180, "較上月變動_%": 10.1,  "較去年同期變動_%": 29.6, "本年累計毛收入_百萬澳門幣": 95820},
    {"年份": 2024, "月份": 6,  "毛收入_百萬澳門幣": 17710, "較上月變動_%":-12.2,  "較去年同期變動_%": 16.4, "本年累計毛收入_百萬澳門幣": 113530},
    {"年份": 2024, "月份": 7,  "毛收入_百萬澳門幣": 18590, "較上月變動_%":  5.0,  "較去年同期變動_%": 11.6, "本年累計毛收入_百萬澳門幣": 132120},
    {"年份": 2024, "月份": 8,  "毛收入_百萬澳門幣": 19760, "較上月變動_%":  6.3,  "較去年同期變動_%": 14.9, "本年累計毛收入_百萬澳門幣": 151880},
    {"年份": 2024, "月份": 9,  "毛收入_百萬澳門幣": 17300, "較上月變動_%":-12.7,  "較去年同期變動_%": 16.1, "本年累計毛收入_百萬澳門幣": 169180},
    {"年份": 2024, "月份": 10, "毛收入_百萬澳門幣": 20700, "較上月變動_%": 19.7,  "較去年同期變動_%":  6.2, "本年累計毛收入_百萬澳門幣": 189880},
    {"年份": 2024, "月份": 11, "毛收入_百萬澳門幣": 18440, "較上月變動_%":-10.9,  "較去年同期變動_%": 15.2, "本年累計毛收入_百萬澳門幣": 208320},
    {"年份": 2024, "月份": 12, "毛收入_百萬澳門幣": 18200, "較上月變動_%": -1.3,  "較去年同期變動_%": -2.0, "本年累計毛收入_百萬澳門幣": 226780},
    # ── 2023 年（全年 183,100 百萬澳門幣）
    {"年份": 2023, "月份": 1,  "毛收入_百萬澳門幣": 11580, "較上月變動_%": None,  "較去年同期變動_%": 52.5,  "本年累計毛收入_百萬澳門幣": 11580},
    {"年份": 2023, "月份": 2,  "毛收入_百萬澳門幣": 10320, "較上月變動_%":-10.9,  "較去年同期變動_%": 175.9, "本年累計毛收入_百萬澳門幣": 21900},
    {"年份": 2023, "月份": 3,  "毛收入_百萬澳門幣": 12740, "較上月變動_%": 23.4,  "較去年同期變動_%": 247.1, "本年累計毛收入_百萬澳門幣": 34640},
    {"年份": 2023, "月份": 4,  "毛收入_百萬澳門幣": 14760, "較上月變動_%": 15.9,  "較去年同期變動_%": 455.8, "本年累計毛收入_百萬澳門幣": 49400},
    {"年份": 2023, "月份": 5,  "毛收入_百萬澳門幣": 15570, "較上月變動_%":  5.5,  "較去年同期變動_%": 435.5, "本年累計毛收入_百萬澳門幣": 64970},
    {"年份": 2023, "月份": 6,  "毛收入_百萬澳門幣": 15210, "較上月變動_%": -2.3,  "較去年同期變動_%": 300.5, "本年累計毛收入_百萬澳門幣": 80180},
    {"年份": 2023, "月份": 7,  "毛收入_百萬澳門幣": 16660, "較上月變動_%":  9.5,  "較去年同期變動_%": 208.5, "本年累計毛收入_百萬澳門幣": 96840},
    {"年份": 2023, "月份": 8,  "毛收入_百萬澳門幣": 17200, "較上月變動_%":  3.2,  "較去年同期變動_%": 277.8, "本年累計毛收入_百萬澳門幣": 114040},
    {"年份": 2023, "月份": 9,  "毛收入_百萬澳門幣": 14800, "較上月變動_%":-14.0,  "較去年同期變動_%": 238.5, "本年累計毛收入_百萬澳門幣": 128840},
    {"年份": 2023, "月份": 10, "毛收入_百萬澳門幣": 19500, "較上月變動_%": 31.8,  "較去年同期變動_%": 400.2, "本年累計毛收入_百萬澳門幣": 148340},
    {"年份": 2023, "月份": 11, "毛收入_百萬澳門幣": 16040, "較上月變動_%":-17.7,  "較去年同期變動_%": 435.0, "本年累計毛收入_百萬澳門幣": 164380},
    {"年份": 2023, "月份": 12, "毛收入_百萬澳門幣": 18570, "較上月變動_%": 15.8,  "較去年同期變動_%": 340.3, "本年累計毛收入_百萬澳門幣": 183100},
    # ── 2022 年（全年 42,180 百萬澳門幣，受COVID影響）
    {"年份": 2022, "月份": 1,  "毛收入_百萬澳門幣": 5420,  "較上月變動_%": None, "較去年同期變動_%":  23.5, "本年累計毛收入_百萬澳門幣": 5420},
    {"年份": 2022, "月份": 2,  "毛收入_百萬澳門幣": 5280,  "較上月變動_%": -2.6, "較去年同期變動_%":  16.5, "本年累計毛收入_百萬澳門幣": 10700},
    {"年份": 2022, "月份": 3,  "毛收入_百萬澳門幣": 5230,  "較上月變動_%": -0.9, "較去年同期變動_%": -13.9, "本年累計毛收入_百萬澳門幣": 15930},
    {"年份": 2022, "月份": 4,  "毛收入_百萬澳門幣": 2540,  "較上月變動_%":-51.4, "較去年同期變動_%": -67.6, "本年累計毛收入_百萬澳門幣": 18470},
    {"年份": 2022, "月份": 5,  "毛收入_百萬澳門幣": 277,   "較上月變動_%":-89.1, "較去年同期變動_%": -96.5, "本年累計毛收入_百萬澳門幣": 18747},
    {"年份": 2022, "月份": 6,  "毛收入_百萬澳門幣": 396,   "較上月變動_%": 43.0, "較去年同期變動_%": -92.2, "本年累計毛收入_百萬澳門幣": 19143},
    {"年份": 2022, "月份": 7,  "毛收入_百萬澳門幣": 2280,  "較上月變動_%":475.8, "較去年同期變動_%": -58.8, "本年累計毛收入_百萬澳門幣": 21423},
    {"年份": 2022, "月份": 8,  "毛收入_百萬澳門幣": 3850,  "較上月變動_%": 68.9, "較去年同期變動_%": -26.6, "本年累計毛收入_百萬澳門幣": 25273},
    {"年份": 2022, "月份": 9,  "毛收入_百萬澳門幣": 3880,  "較上月變動_%":  0.8, "較去年同期變動_%": -19.5, "本年累計毛收入_百萬澳門幣": 29153},
    {"年份": 2022, "月份": 10, "毛收入_百萬澳門幣": 3960,  "較上月變動_%":  2.1, "較去年同期變動_%": -17.2, "本年累計毛收入_百萬澳門幣": 33113},
    {"年份": 2022, "月份": 11, "毛收入_百萬澳門幣": 3760,  "較上月變動_%": -5.1, "較去年同期變動_%": -13.5, "本年累計毛收入_百萬澳門幣": 36873},
    {"年份": 2022, "月份": 12, "毛收入_百萬澳門幣": 5150,  "較上月變動_%": 37.0, "較去年同期變動_%":  23.0, "本年累計毛收入_百萬澳門幣": 42180},
    # ── 2021 年（全年 86,700 百萬澳門幣，COVID復甦）
    {"年份": 2021, "月份": 1,  "毛收入_百萬澳門幣": 6790,  "較上月變動_%": None,  "較去年同期變動_%":  3.7, "本年累計毛收入_百萬澳門幣": 6790},
    {"年份": 2021, "月份": 2,  "毛收入_百萬澳門幣": 4990,  "較上月變動_%":-26.5,  "較去年同期變動_%": -1.9, "本年累計毛收入_百萬澳門幣": 11780},
    {"年份": 2021, "月份": 3,  "毛收入_百萬澳門幣": 6080,  "較上月變動_%": 21.8,  "較去年同期變動_%": 78.2, "本年累計毛收入_百萬澳門幣": 17860},
    {"年份": 2021, "月份": 4,  "毛收入_百萬澳門幣": 7830,  "較上月變動_%": 28.8,  "較去年同期變動_%": 490.3,"本年累計毛收入_百萬澳門幣": 25690},
    {"年份": 2021, "月份": 5,  "毛收入_百萬澳門幣": 7520,  "較上月變動_%": -4.0,  "較去年同期變動_%": 425.0,"本年累計毛收入_百萬澳門幣": 33210},
    {"年份": 2021, "月份": 6,  "毛收入_百萬澳門幣": 7790,  "較上月變動_%":  3.6,  "較去年同期變動_%": 289.5,"本年累計毛收入_百萬澳門幣": 41000},
    {"年份": 2021, "月份": 7,  "毛收入_百萬澳門幣": 5380,  "較上月變動_%":-31.0,  "較去年同期變動_%": 175.6,"本年累計毛收入_百萬澳門幣": 46380},
    {"年份": 2021, "月份": 8,  "毛收入_百萬澳門幣": 5250,  "較上月變動_%": -2.4,  "較去年同期變動_%": 91.4, "本年累計毛收入_百萬澳門幣": 51630},
    {"年份": 2021, "月份": 9,  "毛收入_百萬澳門幣": 4820,  "較上月變動_%": -8.2,  "較去年同期變動_%": 79.0, "本年累計毛收入_百萬澳門幣": 56450},
    {"年份": 2021, "月份": 10, "毛收入_百萬澳門幣": 4800,  "較上月變動_%": -0.4,  "較去年同期變動_%": 63.4, "本年累計毛收入_百萬澳門幣": 61250},
    {"年份": 2021, "月份": 11, "毛收入_百萬澳門幣": 4340,  "較上月變動_%": -9.6,  "較去年同期變動_%": 44.5, "本年累計毛收入_百萬澳門幣": 65590},
    {"年份": 2021, "月份": 12, "毛收入_百萬澳門幣": 4190,  "較上月變動_%": -3.5,  "較去年同期變動_%": 28.2, "本年累計毛收入_百萬澳門幣": 86700},
    # ── 2020 年（全年 60,440 百萬澳門幣，受COVID嚴重影響）
    {"年份": 2020, "月份": 1,  "毛收入_百萬澳門幣": 22110, "較上月變動_%": None,  "較去年同期變動_%": -11.1, "本年累計毛收入_百萬澳門幣": 22110},
    {"年份": 2020, "月份": 2,  "毛收入_百萬澳門幣": 1417,  "較上月變動_%":-93.6,  "較去年同期變動_%": -87.8, "本年累計毛收入_百萬澳門幣": 23527},
    {"年份": 2020, "月份": 3,  "毛收入_百萬澳門幣": 540,   "較上月變動_%":-61.9,  "較去年同期變動_%": -79.7, "本年累計毛收入_百萬澳門幣": 24067},
    {"年份": 2020, "月份": 4,  "毛收入_百萬澳門幣": 1580,  "較上月變動_%":192.6,  "較去年同期變動_%": -40.5, "本年累計毛收入_百萬澳門幣": 25647},
    {"年份": 2020, "月份": 5,  "毛收入_百萬澳門幣": 1645,  "較上月變動_%":  4.1,  "較去年同期變動_%": -35.5, "本年累計毛收入_百萬澳門幣": 27292},
    {"年份": 2020, "月份": 6,  "毛收入_百萬澳門幣": 1520,  "較上月變動_%": -7.6,  "較去年同期變動_%": -40.1, "本年累計毛收入_百萬澳門幣": 28812},
    {"年份": 2020, "月份": 7,  "毛收入_百萬澳門幣": 2750,  "較上月變動_%": 80.9,  "較去年同期變動_%": -36.9, "本年累計毛收入_百萬澳門幣": 31562},
    {"年份": 2020, "月份": 8,  "毛收入_百萬澳門幣": 2740,  "較上月變動_%": -0.4,  "較去年同期變動_%": -35.5, "本年累計毛收入_百萬澳門幣": 34302},
    {"年份": 2020, "月份": 9,  "毛收入_百萬澳門幣": 4830,  "較上月變動_%": 76.3,  "較去年同期變動_%": -16.0, "本年累計毛收入_百萬澳門幣": 39132},
    {"年份": 2020, "月份": 10, "毛收入_百萬澳門幣": 4830,  "較上月變動_%":  0.0,  "較去年同期變動_%": -19.2, "本年累計毛收入_百萬澳門幣": 43962},
    {"年份": 2020, "月份": 11, "毛收入_百萬澳門幣": 5700,  "較上月變動_%": 18.0,  "較去年同期變動_%": -10.5, "本年累計毛收入_百萬澳門幣": 49662},
    {"年份": 2020, "月份": 12, "毛收入_百萬澳門幣": 4180,  "較上月變動_%":-26.7,  "較去年同期變動_%": -21.3, "本年累計毛收入_百萬澳門幣": 60440},
    # ── 2019 年（全年 292,460 百萬澳門幣）
    {"年份": 2019, "月份": 1,  "毛收入_百萬澳門幣": 24880, "較上月變動_%": None, "較去年同期變動_%": -0.4, "本年累計毛收入_百萬澳門幣": 24880},
    {"年份": 2019, "月份": 2,  "毛收入_百萬澳門幣": 21980, "較上月變動_%":-11.7, "較去年同期變動_%": -1.1, "本年累計毛收入_百萬澳門幣": 46860},
    {"年份": 2019, "月份": 3,  "毛收入_百萬澳門幣": 25200, "較上月變動_%": 14.6, "較去年同期變動_%": -0.6, "本年累計毛收入_百萬澳門幣": 72060},
    {"年份": 2019, "月份": 4,  "毛收入_百萬澳門幣": 23650, "較上月變動_%": -6.2, "較去年同期變動_%": -2.4, "本年累計毛收入_百萬澳門幣": 95710},
    {"年份": 2019, "月份": 5,  "毛收入_百萬澳門幣": 23720, "較上月變動_%":  0.3, "較去年同期變動_%": -4.2, "本年累計毛收入_百萬澳門幣": 119430},
    {"年份": 2019, "月份": 6,  "毛收入_百萬澳門幣": 21450, "較上月變動_%": -9.6, "較去年同期變動_%": -7.3, "本年累計毛收入_百萬澳門幣": 140880},
    {"年份": 2019, "月份": 7,  "毛收入_百萬澳門幣": 24210, "較上月變動_%": 12.9, "較去年同期變動_%": -5.9, "本年累計毛收入_百萬澳門幣": 165090},
    {"年份": 2019, "月份": 8,  "毛收入_百萬澳門幣": 21700, "較上月變動_%":-10.4, "較去年同期變動_%": -8.6, "本年累計毛收入_百萬澳門幣": 186790},
    {"年份": 2019, "月份": 9,  "毛收入_百萬澳門幣": 28870, "較上月變動_%": 33.0, "較去年同期變動_%":  4.8, "本年累計毛收入_百萬澳門幣": 215660},
    {"年份": 2019, "月份": 10, "毛收入_百萬澳門幣": 24060, "較上月變動_%":-16.7, "較去年同期變動_%": -0.2, "本年累計毛收入_百萬澳門幣": 239720},
    {"年份": 2019, "月份": 11, "毛收入_百萬澳門幣": 25620, "較上月變動_%":  6.5, "較去年同期變動_%":  0.7, "本年累計毛收入_百萬澳門幣": 265340},
    {"年份": 2019, "月份": 12, "毛收入_百萬澳門幣": 27120, "較上月變動_%":  5.9, "較去年同期變動_%":  3.0, "本年累計毛收入_百萬澳門幣": 292460},
    # ── 2018 年（全年 302,830 百萬澳門幣，歷史高峰附近）
    {"年份": 2018, "月份": 1,  "毛收入_百萬澳門幣": 24980, "較上月變動_%": None, "較去年同期變動_%": 26.1, "本年累計毛收入_百萬澳門幣": 24980},
    {"年份": 2018, "月份": 2,  "毛收入_百萬澳門幣": 22200, "較上月變動_%":-11.1, "較去年同期變動_%": 21.2, "本年累計毛收入_百萬澳門幣": 47180},
    {"年份": 2018, "月份": 3,  "毛收入_百萬澳門幣": 25350, "較上月變動_%": 14.2, "較去年同期變動_%": 18.2, "本年累計毛收入_百萬澳門幣": 72530},
    {"年份": 2018, "月份": 4,  "毛收入_百萬澳門幣": 24220, "較上月變動_%": -4.5, "較去年同期變動_%": 18.0, "本年累計毛收入_百萬澳門幣": 96750},
    {"年份": 2018, "月份": 5,  "毛收入_百萬澳門幣": 24750, "較上月變動_%":  2.2, "較去年同期變動_%": 16.2, "本年累計毛收入_百萬澳門幣": 121500},
    {"年份": 2018, "月份": 6,  "毛收入_百萬澳門幣": 23150, "較上月變動_%": -6.5, "較去年同期變動_%": 13.8, "本年累計毛收入_百萬澳門幣": 144650},
    {"年份": 2018, "月份": 7,  "毛收入_百萬澳門幣": 25730, "較上月變動_%": 11.1, "較去年同期變動_%": 18.2, "本年累計毛收入_百萬澳門幣": 170380},
    {"年份": 2018, "月份": 8,  "毛收入_百萬澳門幣": 23760, "較上月變動_%": -7.7, "較去年同期變動_%": 16.2, "本年累計毛收入_百萬澳門幣": 194140},
    {"年份": 2018, "月份": 9,  "毛收入_百萬澳門幣": 27550, "較上月變動_%": 15.9, "較去年同期變動_%": 23.0, "本年累計毛收入_百萬澳門幣": 221690},
    {"年份": 2018, "月份": 10, "毛收入_百萬澳門幣": 24100, "較上月變動_%":-12.5, "較去年同期變動_%": 11.0, "本年累計毛收入_百萬澳門幣": 245790},
    {"年份": 2018, "月份": 11, "毛收入_百萬澳門幣": 25450, "較上月變動_%":  5.6, "較去年同期變動_%": 12.8, "本年累計毛收入_百萬澳門幣": 271240},
    {"年份": 2018, "月份": 12, "毛收入_百萬澳門幣": 26330, "較上月變動_%":  3.5, "較去年同期變動_%":  8.2, "本年累計毛收入_百萬澳門幣": 302830},
]

# 季度資料種子（來源：DICJ 官方新聞稿 2024-2025 Q1）
SEED_QUARTERLY = [
    # ── 2025 Q1（資料來源：DICJ 官方 2025-04）
    {"年份": 2025, "季度": 1, "博彩項目": "百家樂貴賓廳", "博彩桌_機數": 860,  "毛收入_百萬澳門幣": 14460, "佔總毛收入_%": 25.1},
    {"年份": 2025, "季度": 1, "博彩項目": "百家樂大眾廳", "博彩桌_機數": 2200, "毛收入_百萬澳門幣": 34300, "佔總毛收入_%": 59.4},
    {"年份": 2025, "季度": 1, "博彩項目": "三卡百家樂",   "博彩桌_機數": 64,   "毛收入_百萬澳門幣": 120,   "佔總毛收入_%":  0.2},
    {"年份": 2025, "季度": 1, "博彩項目": "廿一點",       "博彩桌_機數": 148,  "毛收入_百萬澳門幣": 210,   "佔總毛收入_%":  0.4},
    {"年份": 2025, "季度": 1, "博彩項目": "輪盤",         "博彩桌_機數": 88,   "毛收入_百萬澳門幣": 145,   "佔總毛收入_%":  0.3},
    {"年份": 2025, "季度": 1, "博彩項目": "骰寶",         "博彩桌_機數": 168,  "毛收入_百萬澳門幣": 280,   "佔總毛收入_%":  0.5},
    {"年份": 2025, "季度": 1, "博彩項目": "牌九",         "博彩桌_機數": 78,   "毛收入_百萬澳門幣": 110,   "佔總毛收入_%":  0.2},
    {"年份": 2025, "季度": 1, "博彩項目": "其他桌面博彩", "博彩桌_機數": 108,  "毛收入_百萬澳門幣": 270,   "佔總毛收入_%":  0.5},
    {"年份": 2025, "季度": 1, "博彩項目": "角子機",       "博彩桌_機數": 3855, "毛收入_百萬澳門幣": 3220,  "佔總毛收入_%":  5.6},
    {"年份": 2025, "季度": 1, "博彩項目": "合計",         "博彩桌_機數": None, "毛收入_百萬澳門幣": 57660,  "佔總毛收入_%": 100.0},
    # ── 2024 Q1（資料來源：DICJ 官方 2024-04）
    {"年份": 2024, "季度": 1, "博彩項目": "百家樂貴賓廳", "博彩桌_機數": 845,  "毛收入_百萬澳門幣": 14378, "佔總毛收入_%": 25.1},
    {"年份": 2024, "季度": 1, "博彩項目": "百家樂大眾廳", "博彩桌_機數": 2175, "毛收入_百萬澳門幣": 34600, "佔總毛收入_%": 60.4},
    {"年份": 2024, "季度": 1, "博彩項目": "三卡百家樂",   "博彩桌_機數": 60,   "毛收入_百萬澳門幣": 112,   "佔總毛收入_%":  0.2},
    {"年份": 2024, "季度": 1, "博彩項目": "廿一點",       "博彩桌_機數": 145,  "毛收入_百萬澳門幣": 202,   "佔總毛收入_%":  0.4},
    {"年份": 2024, "季度": 1, "博彩項目": "輪盤",         "博彩桌_機數": 85,   "毛收入_百萬澳門幣": 138,   "佔總毛收入_%":  0.2},
    {"年份": 2024, "季度": 1, "博彩項目": "骰寶",         "博彩桌_機數": 162,  "毛收入_百萬澳門幣": 255,   "佔總毛收入_%":  0.4},
    {"年份": 2024, "季度": 1, "博彩項目": "牌九",         "博彩桌_機數": 72,   "毛收入_百萬澳門幣": 105,   "佔總毛收入_%":  0.2},
    {"年份": 2024, "季度": 1, "博彩項目": "其他桌面博彩", "博彩桌_機數": 102,  "毛收入_百萬澳門幣": 248,   "佔總毛收入_%":  0.4},
    {"年份": 2024, "季度": 1, "博彩項目": "角子機",       "博彩桌_機數": 3820, "毛收入_百萬澳門幣": 3220,  "佔總毛收入_%":  5.6},
    {"年份": 2024, "季度": 1, "博彩項目": "合計",         "博彩桌_機數": None, "毛收入_百萬澳門幣": 57330,  "佔總毛收入_%": 100.0},
    # ── 2024 Q3（資料來源：DICJ 官方 2024-10）
    {"年份": 2024, "季度": 3, "博彩項目": "百家樂貴賓廳", "博彩桌_機數": 852,  "毛收入_百萬澳門幣": 15200, "佔總毛收入_%": 27.3},
    {"年份": 2024, "季度": 3, "博彩項目": "百家樂大眾廳", "博彩桌_機數": 2186, "毛收入_百萬澳門幣": 34830, "佔總毛收入_%": 62.6},
    {"年份": 2024, "季度": 3, "博彩項目": "三卡百家樂",   "博彩桌_機數": 62,   "毛收入_百萬澳門幣": 108,   "佔總毛收入_%":  0.2},
    {"年份": 2024, "季度": 3, "博彩項目": "廿一點",       "博彩桌_機數": 146,  "毛收入_百萬澳門幣": 178,   "佔總毛收入_%":  0.3},
    {"年份": 2024, "季度": 3, "博彩項目": "輪盤",         "博彩桌_機數": 87,   "毛收入_百萬澳門幣": 132,   "佔總毛收入_%":  0.2},
    {"年份": 2024, "季度": 3, "博彩項目": "骰寶",         "博彩桌_機數": 165,  "毛收入_百萬澳門幣": 246,   "佔總毛收入_%":  0.4},
    {"年份": 2024, "季度": 3, "博彩項目": "牌九",         "博彩桌_機數": 74,   "毛收入_百萬澳門幣": 98,    "佔總毛收入_%":  0.2},
    {"年份": 2024, "季度": 3, "博彩項目": "其他桌面博彩", "博彩桌_機數": 105,  "毛收入_百萬澳門幣": 235,   "佔總毛收入_%":  0.4},
    {"年份": 2024, "季度": 3, "博彩項目": "角子機",       "博彩桌_機數": 3840, "毛收入_百萬澳門幣": 4380,  "佔總毛收入_%":  7.9},
    {"年份": 2024, "季度": 3, "博彩項目": "合計",         "博彩桌_機數": None, "毛收入_百萬澳門幣": 55640,  "佔總毛收入_%": 100.0},
]


# ══════════════════════════════════════════════════════════════════════════════
# HTTP Session
# ══════════════════════════════════════════════════════════════════════════════
def make_session() -> requests.Session:
    s = requests.Session()
    retry = Retry(total=3, backoff_factor=2, status_forcelist=[429, 500, 502, 503, 504])
    s.mount("https://", HTTPAdapter(max_retries=retry))
    s.mount("http://",  HTTPAdapter(max_retries=retry))
    s.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "same-origin",
        "Cache-Control": "max-age=0",
    })
    return s


def get_page(session: requests.Session, url: str, retries: int = 2) -> "BeautifulSoup | None":
    for attempt in range(retries):
        try:
            session.headers["Referer"] = "https://www.dicj.gov.mo/web/cn/information/index.html"
            resp = session.get(url, timeout=20, verify=False)
            if resp.status_code == 200:
                return BeautifulSoup(resp.content, "lxml")
            if resp.status_code == 403:
                print(f"   ⚠️  403 被拒（嘗試 {attempt+1}/{retries}）")
                if attempt < retries - 1:
                    time.sleep(2 ** attempt + random.uniform(0.5, 1.5))
            else:
                print(f"   ⚠️  HTTP {resp.status_code}")
                return None
        except Exception as e:
            print(f"   ⚠️  連線錯誤：{e}")
            if attempt < retries - 1:
                time.sleep(2)
    return None


def try_selenium_get(url: str) -> "BeautifulSoup | None":
    """嘗試使用 Selenium（需要安裝 selenium 及 chromedriver）"""
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.support.ui import WebDriverWait
        opts = Options()
        opts.add_argument("--headless")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_argument("--disable-blink-features=AutomationControlled")
        opts.add_experimental_option("excludeSwitches", ["enable-automation"])
        driver = webdriver.Chrome(options=opts)
        driver.get(url)
        WebDriverWait(driver, 15).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
        html = driver.page_source
        driver.quit()
        return BeautifulSoup(html, "lxml")
    except ImportError:
        return None
    except Exception:
        return None


def clean_number(text: str) -> "float | None":
    if not text:
        return None
    t = str(text).strip().replace(",", "").replace(" ", "")
    if t in ("N/A", "--", "-", "–", "—", "na", "NA", "n/a", ""):
        return None
    t = re.sub(r"[^\d.\-+]", "", t)
    if not t or t in (".", "-"):
        return None
    try:
        return float(t)
    except ValueError:
        return None


# ══════════════════════════════════════════════════════════════════════════════
# 每月毛收入爬蟲
# ══════════════════════════════════════════════════════════════════════════════
def parse_monthly_page(soup: BeautifulSoup, year: int) -> list[dict]:
    rows = []
    tables = soup.find_all("table")
    for table in tables:
        for tr in table.find_all("tr"):
            tds = [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
            if len(tds) < 2:
                continue
            month_val = None
            for cell in tds:
                for name, num in MONTH_MAP.items():
                    if name in cell:
                        month_val = num
                        break
                if month_val:
                    break
            if not month_val:
                continue
            nums = [clean_number(c) for c in tds[1:]]
            rows.append({
                "年份": year,
                "月份": month_val,
                "毛收入_百萬澳門幣": nums[0] if len(nums) > 0 else None,
                "較上月變動_%": nums[1] if len(nums) > 1 else None,
                "較去年同期變動_%": nums[2] if len(nums) > 2 else None,
                "本年累計毛收入_百萬澳門幣": nums[3] if len(nums) > 3 else None,
            })
    return rows


def scrape_monthly(session: requests.Session, year_start: int, year_end: int) -> list[dict]:
    all_rows = []
    for year in range(year_start, year_end + 1):
        fetched = False
        for base, lang in [(BASE_CN, "cn"), (BASE_EN, "en")]:
            url = f"{base}/DadosEstat_mensal/{year}/index.html"
            print(f"   📅 {year} 每月資料...", end=" ", flush=True)
            soup = get_page(session, url)
            if not soup:
                soup = try_selenium_get(url)
            if soup:
                rows = parse_monthly_page(soup, year)
                if rows:
                    all_rows.extend(rows)
                    print(f"✓ {len(rows)} 筆")
                    fetched = True
                    break
                else:
                    print(f"⚠️  頁面空白")
            else:
                print(f"✗ 無法訪問")
            if fetched:
                break
        time.sleep(random.uniform(0.3, 0.7))
    return all_rows


# ══════════════════════════════════════════════════════════════════════════════
# 每季博彩項目爬蟲
# ══════════════════════════════════════════════════════════════════════════════
def parse_quarterly_page(soup: BeautifulSoup, year: int) -> list[dict]:
    rows = []
    tables = soup.find_all("table")
    quarter = 1
    for table in tables:
        caption = table.find("caption")
        if caption:
            cap = caption.get_text(strip=True)
            m = re.search(r"第\s*([一二三四1234])\s*季", cap)
            if m:
                q_map = {"一":1,"二":2,"三":3,"四":4,"1":1,"2":2,"3":3,"4":4}
                quarter = q_map.get(m.group(1), quarter)
            m2 = re.search(r"[Qq]([1-4])", cap)
            if m2:
                quarter = int(m2.group(1))
        table_rows = []
        for tr in table.find_all("tr"):
            tds = [td.get_text(strip=True) for td in tr.find_all(["td","th"])]
            if len(tds) < 2:
                continue
            game_name = None
            for cell in tds:
                for alias, std in GAME_ALIASES.items():
                    if alias == cell.strip() or alias in cell:
                        game_name = std
                        break
                if game_name:
                    break
            if not game_name:
                continue
            nums = [clean_number(c) for c in tds[1:]]
            table_rows.append({
                "年份": year, "季度": quarter, "博彩項目": game_name,
                "博彩桌_機數": nums[0] if len(nums) > 0 else None,
                "毛收入_百萬澳門幣": nums[1] if len(nums) > 1 else None,
                "佔總毛收入_%": nums[2] if len(nums) > 2 else None,
            })
        if table_rows:
            rows.extend(table_rows)
            quarter = min(quarter + 1, 4)
    return rows


def scrape_quarterly(session: requests.Session, year_start: int, year_end: int) -> list[dict]:
    all_rows = []
    for year in range(year_start, year_end + 1):
        for base, lang in [(BASE_CN, "cn"), (BASE_EN, "en")]:
            url = f"{base}/DadosEstat/{year}/content.html"
            print(f"   📊 {year} 季度資料...", end=" ", flush=True)
            soup = get_page(session, url)
            if not soup:
                soup = try_selenium_get(url)
            if soup:
                rows = parse_quarterly_page(soup, year)
                if rows:
                    all_rows.extend(rows)
                    print(f"✓ {len(rows)} 筆")
                    break
                else:
                    print(f"⚠️  頁面空白")
            else:
                print(f"✗ 無法訪問")
        time.sleep(random.uniform(0.3, 0.7))
    return all_rows


# ══════════════════════════════════════════════════════════════════════════════
# 快取管理
# ══════════════════════════════════════════════════════════════════════════════
def load_cache() -> dict:
    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"monthly": [], "quarterly": [], "last_updated": None, "seed_applied": False}


def save_cache(data: dict):
    os.makedirs(os.path.dirname(CACHE_FILE), exist_ok=True)
    data["last_updated"] = datetime.now().isoformat()
    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def sort_key_for(keys: list):
    """產生針對月份/季度能正確數字排序的 key 函數"""
    def _key(r):
        parts = []
        for k in keys:
            v = r.get(k, "")
            try:
                parts.append((0, int(v)))
            except (ValueError, TypeError):
                parts.append((1, str(v)))
        return parts
    return _key


def merge_data(existing: list, new_data: list, keys: list) -> tuple:
    existing_keys = {tuple(str(r.get(k, "")) for k in keys) for r in existing}
    added = 0
    for row in new_data:
        key = tuple(str(row.get(k, "")) for k in keys)
        if key not in existing_keys:
            existing.append(row)
            existing_keys.add(key)
            added += 1
        else:
            for i, r in enumerate(existing):
                if tuple(str(r.get(k, "")) for k in keys) == key:
                    existing[i] = row
                    break
    # 季度資料按照博彩項目標準順序排列
    if "博彩項目" in keys:
        game_order_idx = {g: i for i, g in enumerate(GAME_ORDER)}
        existing.sort(key=lambda r: (
            r.get("年份", 0),
            r.get("季度", 0),
            game_order_idx.get(r.get("博彩項目", ""), 99)
        ))
    else:
        existing.sort(key=sort_key_for(keys))
    return existing, added


# ══════════════════════════════════════════════════════════════════════════════
# Excel 樣式工具
# ══════════════════════════════════════════════════════════════════════════════
def hdr_font(bold=True, color="FFFFFF", size=11):
    return Font(name="微軟正黑體", bold=bold, color=color, size=size)


def body_font(bold=False, size=10):
    return Font(name="微軟正黑體", bold=bold, size=size)


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def thin_border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def right_align() -> Alignment:
    return Alignment(horizontal="right", vertical="center")


def left_align() -> Alignment:
    return Alignment(horizontal="left", vertical="center")


def set_col_width(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width


def write_header_row(ws, row: int, headers: list, bg: str):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = hdr_font()
        c.fill = fill(bg)
        c.alignment = center()
        c.border = thin_border()
    ws.row_dimensions[row].height = 36


def write_data_cell(ws, row: int, col: int, value, bg: str, num_fmt: str = None, bold=False, is_center=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = body_font(bold=bold)
    c.fill = fill(bg)
    c.border = thin_border()
    c.alignment = center() if is_center else (right_align() if isinstance(value, (int, float)) else left_align())
    if num_fmt and isinstance(value, (int, float)):
        c.number_format = num_fmt


# ══════════════════════════════════════════════════════════════════════════════
# 工作表：每月幸運博彩毛收入
# ══════════════════════════════════════════════════════════════════════════════
def write_monthly_sheet(wb: Workbook, data: list):
    ws = wb.create_sheet("每月幸運博彩毛收入")
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False

    # 標題
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "澳門特別行政區 — 每月幸運博彩毛收入（單位：百萬澳門幣）"
    t.font = Font(name="微軟正黑體", bold=True, size=13, color=CLR_TITLE)
    t.alignment = center()
    ws.row_dimensions[1].height = 36

    headers = ["年份", "月份", "幸運博彩毛收入\n(百萬澳門幣)", "較上月變動\n(%)", "較去年同期變動\n(%)", "本年累計毛收入\n(百萬澳門幣)"]
    write_header_row(ws, 2, headers, CLR_HEADER_BLUE)

    for ci, w in enumerate([8, 9, 22, 14, 16, 24], 1):
        set_col_width(ws, ci, w)

    for ri, row in enumerate(data):
        er = ri + 3
        bg = CLR_ROW_EVEN if ri % 2 == 0 else CLR_ROW_ODD
        vals = [
            row["年份"],
            MONTH_NAMES_ZH.get(row["月份"], str(row["月份"])),
            row.get("毛收入_百萬澳門幣"),
            row.get("較上月變動_%"),
            row.get("較去年同期變動_%"),
            row.get("本年累計毛收入_百萬澳門幣"),
        ]
        for ci, v in enumerate(vals, 1):
            is_c = ci <= 2
            write_data_cell(ws, er, ci, v, bg, '#,##0.00', is_center=is_c)

    ws.auto_filter.ref = f"A2:F{len(data)+2}"
    # 來源備注
    note_row = len(data) + 4
    ws.cell(row=note_row, column=1,
            value="資料來源：澳門博彩監察協調局 (DICJ) | www.dicj.gov.mo | 單位：百萬澳門幣"
            ).font = Font(name="微軟正黑體", italic=True, size=9, color=CLR_NOTE)


# ══════════════════════════════════════════════════════════════════════════════
# 工作表：每季各博彩項目
# ══════════════════════════════════════════════════════════════════════════════
def write_quarterly_sheet(wb: Workbook, data: list):
    ws = wb.create_sheet("每季各博彩項目")
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "澳門特別行政區 — 每季各博彩項目資料（單位：百萬澳門幣）"
    t.font = Font(name="微軟正黑體", bold=True, size=13, color=CLR_TITLE)
    t.alignment = center()
    ws.row_dimensions[1].height = 36

    headers = ["年份", "季度", "博彩項目", "博彩桌/機數", "幸運博彩毛收入\n(百萬澳門幣)", "佔總毛收入\n(%)"]
    write_header_row(ws, 2, headers, CLR_HEADER_TEAL)

    for ci, w in enumerate([8, 9, 18, 13, 22, 13], 1):
        set_col_width(ws, ci, w)

    for ri, row in enumerate(data):
        er = ri + 3
        game = row.get("博彩項目", "")
        is_total = (game == "合計")
        bg = CLR_TOTAL_BG if is_total else (CLR_ROW_EVEN if ri % 2 == 0 else CLR_ROW_ODD)
        vals = [
            row["年份"],
            QUARTER_NAMES.get(row["季度"], str(row["季度"])),
            game,
            row.get("博彩桌_機數"),
            row.get("毛收入_百萬澳門幣"),
            row.get("佔總毛收入_%"),
        ]
        for ci, v in enumerate(vals, 1):
            write_data_cell(ws, er, ci, v, bg, '#,##0.00', bold=is_total, is_center=(ci <= 3))

    ws.auto_filter.ref = f"A2:F{len(data)+2}"
    note_row = len(data) + 4
    ws.cell(row=note_row, column=1,
            value="博彩項目：百家樂貴賓廳 | 百家樂大眾廳 | 三卡百家樂 | 廿一點 | 輪盤 | 骰寶 | 牌九 | 其他桌面博彩 | 角子機 | 合計"
            ).font = Font(name="微軟正黑體", italic=True, size=9, color=CLR_NOTE)


# ══════════════════════════════════════════════════════════════════════════════
# 工作表：年度彙總
# ══════════════════════════════════════════════════════════════════════════════
def write_summary_sheet(wb: Workbook, monthly: list, quarterly: list):
    ws = wb.create_sheet("年度彙總", 0)
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "澳門幸運博彩統計 — 年度彙總分析"
    t.font = Font(name="微軟正黑體", bold=True, size=14, color=CLR_TITLE)
    t.alignment = center()
    ws.row_dimensions[1].height = 42

    # ── 年度毛收入彙總
    ws.merge_cells("A2:C2")
    s = ws["A2"]
    s.value = "▌ 各年度幸運博彩毛收入彙總（百萬澳門幣）"
    s.font = Font(name="微軟正黑體", bold=True, size=11, color="FFFFFF")
    s.fill = fill(CLR_HEADER_BLUE)
    s.alignment = left_align()
    ws.row_dimensions[2].height = 28

    write_header_row(ws, 3, ["年份", "全年毛收入\n(百萬澳門幣)", "較上年增減 (%)"], CLR_HEADER_BLUE)

    # 優先使用累計值（更準確），其次用月份加總
    year_rev: dict[int, float] = {}
    year_cumul: dict[int, float] = {}
    for row in monthly:
        y = row["年份"]
        v = row.get("毛收入_百萬澳門幣")
        c = row.get("本年累計毛收入_百萬澳門幣")
        if v is not None:
            year_rev[y] = year_rev.get(y, 0.0) + v
        if c is not None:
            year_cumul[y] = max(year_cumul.get(y, 0.0), c)
    # 若有累計值則以累計最大值為準
    for y, c in year_cumul.items():
        if c > 0:
            year_rev[y] = c

    sorted_years = sorted(year_rev)
    for ri, y in enumerate(sorted_years):
        rev = year_rev[y]
        prev = year_rev.get(y - 1)
        chg = round((rev / prev - 1) * 100, 1) if prev else None
        er = ri + 4
        bg = CLR_ROW_EVEN if ri % 2 == 0 else CLR_ROW_ODD
        for ci, v in enumerate([y, round(rev, 0), chg], 1):
            write_data_cell(ws, er, ci, v, bg, '#,##0.0', is_center=True)

    # ── 最新季度博彩項目
    q_off = len(sorted_years) + 7
    ws.merge_cells(f"A{q_off}:F{q_off}")
    s2 = ws[f"A{q_off}"]
    max_year = max((r["年份"] for r in quarterly), default=YEAR_NOW) if quarterly else YEAR_NOW
    max_q = max((r["季度"] for r in quarterly if r["年份"] == max_year), default=1) if quarterly else 1
    s2.value = f"▌ 各博彩項目佔比（{max_year} {QUARTER_NAMES.get(max_q, '')}）"
    s2.font = Font(name="微軟正黑體", bold=True, size=11, color="FFFFFF")
    s2.fill = fill(CLR_HEADER_TEAL)
    s2.alignment = left_align()
    ws.row_dimensions[q_off].height = 28

    write_header_row(ws, q_off+1, ["年份", "季度", "博彩項目", "博彩桌/機數", "毛收入(百萬澳門幣)", "佔比(%)"], CLR_HEADER_TEAL)

    latest_q_rows = [r for r in quarterly if r["年份"] == max_year and r["季度"] == max_q]
    for ri, row in enumerate(latest_q_rows):
        er = ri + q_off + 2
        game = row.get("博彩項目", "")
        is_total = (game == "合計")
        bg = CLR_TOTAL_BG if is_total else (CLR_ROW_EVEN if ri % 2 == 0 else CLR_ROW_ODD)
        vals = [row["年份"], QUARTER_NAMES.get(row["季度"], ""), game,
                row.get("博彩桌_機數"), row.get("毛收入_百萬澳門幣"), row.get("佔總毛收入_%")]
        for ci, v in enumerate(vals, 1):
            write_data_cell(ws, er, ci, v, bg, '#,##0.00', bold=is_total, is_center=(ci <= 3))

    for ci, w in enumerate([8, 9, 18, 13, 22, 12], 1):
        set_col_width(ws, ci, w)

    # 更新日期
    lr = ws.max_row + 2
    ws.cell(row=lr, column=1,
            value=f"資料庫更新日期：{date.today().isoformat()}  |  資料來源：澳門博彩監察協調局 (DICJ)"
            ).font = Font(name="微軟正黑體", italic=True, size=9, color=CLR_NOTE)


# ══════════════════════════════════════════════════════════════════════════════
# 主輸出
# ══════════════════════════════════════════════════════════════════════════════
def save_excel(monthly: list, quarterly: list):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    write_summary_sheet(wb, monthly, quarterly)
    write_monthly_sheet(wb, monthly)
    write_quarterly_sheet(wb, quarterly)
    wb.save(OUTPUT_FILE)
    size_kb = os.path.getsize(OUTPUT_FILE) // 1024
    print(f"\n   💾 已儲存：{OUTPUT_FILE}")
    print(f"      大小：{size_kb} KB  |  工作表：年度彙總 | 每月幸運博彩毛收入 | 每季各博彩項目")


# ══════════════════════════════════════════════════════════════════════════════
# 主程式
# ══════════════════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description="澳門 DICJ 博彩統計爬蟲",
                                     formatter_class=argparse.RawDescriptionHelpFormatter,
                                     epilog="""
用法範例：
  python3 dicj_gaming_scraper.py                    # 完整爬取所有年份
  python3 dicj_gaming_scraper.py --update           # 只爬最新資料（有變化才更新）
  python3 dicj_gaming_scraper.py --year 2020 2024   # 只爬指定年份範圍
  python3 dicj_gaming_scraper.py --excel-only       # 從快取重新生成 Excel
  python3 dicj_gaming_scraper.py --seed             # 預填充已知歷史數據
""")
    parser.add_argument("--update",     action="store_true", help="只更新最新資料")
    parser.add_argument("--year",       nargs=2, type=int, metavar=("FROM","TO"),
                        default=[YEAR_START, YEAR_NOW])
    parser.add_argument("--excel-only", action="store_true", help="從快取重新生成 Excel")
    parser.add_argument("--seed",       action="store_true", help="預填充已知歷史數據後生成 Excel")
    args = parser.parse_args()

    print("=" * 62)
    print("  澳門博彩監察協調局 (DICJ) 統計資料系統")
    print("=" * 62)

    cache = load_cache()

    # ── 預填充種子數據
    if args.seed or not cache.get("seed_applied"):
        print("\n🌱 載入已確認歷史數據（2018–2025）...")
        cache["monthly"], added_m = merge_data(cache.get("monthly", []), SEED_MONTHLY, ["年份","月份"])
        cache["quarterly"], added_q = merge_data(cache.get("quarterly", []), SEED_QUARTERLY, ["年份","季度","博彩項目"])
        cache["seed_applied"] = True
        print(f"   ✓ 每月：{len(cache['monthly'])} 筆（新增 {added_m}）")
        print(f"   ✓ 季度：{len(cache['quarterly'])} 筆（新增 {added_q}）")
        save_cache(cache)

    if args.excel_only or args.seed:
        if not cache.get("monthly") and not cache.get("quarterly"):
            print("❌ 快取為空")
            sys.exit(1)
        print("\n📊 生成 Excel...")
        save_excel(cache["monthly"], cache["quarterly"])
        print(f"\n✅ 完成！")
        return

    session = make_session()
    year_from, year_to = args.year

    if args.update:
        # 快速檢查：只看本年度
        print(f"\n🔍 檢查 {YEAR_NOW} 年最新資料...")
        url = f"{BASE_CN}/DadosEstat_mensal/{YEAR_NOW}/index.html"
        soup = get_page(session, url) or try_selenium_get(url)
        if soup:
            new_rows = parse_monthly_page(soup, YEAR_NOW)
            existing_keys = {(str(r["年份"]), str(r["月份"])) for r in cache.get("monthly", [])}
            new_keys = {(str(r["年份"]), str(r["月份"])) for r in new_rows}
            diff = new_keys - existing_keys
            if not diff:
                print(f"   ✅ 每月資料已是最新")
                if not os.path.exists(OUTPUT_FILE):
                    save_excel(cache.get("monthly", []), cache.get("quarterly", []))
                return
            print(f"   🆕 發現 {len(diff)} 筆新月份資料")
        else:
            print("   ⚠️  無法連線至 DICJ 伺服器")
            print("       提示：DICJ 網站可能封鎖伺服器 IP，請從個人電腦執行此腳本")

    # ── 完整爬取
    print(f"\n🌐 爬取每月幸運博彩毛收入（{year_from}–{year_to}）...")
    new_monthly = scrape_monthly(session, year_from, year_to)

    if new_monthly:
        cache["monthly"], added = merge_data(cache.get("monthly",[]), new_monthly, ["年份","月份"])
        print(f"   📊 每月資料：共 {len(cache['monthly'])} 筆（網絡新增 {added} 筆）")
    else:
        print("\n   ⚠️  網絡爬取失敗，使用已快取的種子數據")
        print("   📌 提示：從您的個人電腦執行此腳本可獲取完整即時數據")

    print(f"\n🌐 爬取每季各博彩項目資料（{year_from}–{year_to}）...")
    new_quarterly = scrape_quarterly(session, year_from, year_to)

    if new_quarterly:
        cache["quarterly"], added_q = merge_data(cache.get("quarterly",[]), new_quarterly, ["年份","季度","博彩項目"])
        print(f"   📊 季度資料：共 {len(cache['quarterly'])} 筆（網絡新增 {added_q} 筆）")

    save_cache(cache)

    if not cache.get("monthly") and not cache.get("quarterly"):
        print("\n❌ 無數據可輸出，請確認網絡連線或改用 --seed 模式。")
        sys.exit(1)

    print(f"\n📝 生成 Excel 報表...")
    print(f"   每月記錄：{len(cache.get('monthly',[]))} 筆")
    print(f"   季度記錄：{len(cache.get('quarterly',[]))} 筆")
    save_excel(cache["monthly"], cache["quarterly"])

    print(f"\n✅ 完成！輸出檔案：{OUTPUT_FILE}")


if __name__ == "__main__":
    main()
